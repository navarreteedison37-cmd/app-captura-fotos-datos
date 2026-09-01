from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

class ExcelGenerator:
    """Genera archivos Excel con datos"""
    
    def __init__(self):
        self.output_dir = 'reportes'
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def generar(self, registros, tipo='todos'):
        """Genera archivo Excel con dos hojas"""
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Eliminar hoja por defecto
            
            # Crear dos hojas
            ws_laboral = wb.create_sheet("Inspección Laboral")
            ws_ingresos = wb.create_sheet("Control de Ingresos")
            
            # Llenar hojas
            self._llenar_hoja_laboral(ws_laboral, registros)
            self._llenar_hoja_ingresos(ws_ingresos, registros)
            
            # Guardar archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            archivo = os.path.join(self.output_dir, f'reporte_{timestamp}.xlsx')
            wb.save(archivo)
            
            return archivo
        except Exception as e:
            raise Exception(f"Error generando Excel: {str(e)}")
    
    def _llenar_hoja_laboral(self, ws, registros):
        """Llena la hoja de Inspección Laboral"""
        # Headers
        headers = [
            "ID", "Número Acta", "RUC", "Razón Social", 
            "Número Carta", "Recepción", "Vínculo", "Fecha", "Hora", "Fecha Creación"
        ]
        
        self._aplicar_estilo_headers(ws, headers)
        
        # Datos
        fila = 2
        for reg in registros:
            if reg['tipo_informe'] == 'inspeccion_laboral':
                ws[f'A{fila}'] = reg['id']
                ws[f'B{fila}'] = reg['documento'].get('numero_acta', '')
                ws[f'C{fila}'] = reg['documento'].get('ruc', '')
                ws[f'D{fila}'] = reg['documento'].get('nombre_razon_social', '')
                ws[f'E{fila}'] = reg['comprobante'].get('numero_carta', '')
                ws[f'F{fila}'] = reg['comprobante'].get('recepcion', '')
                ws[f'G{fila}'] = reg['comprobante'].get('vinculo', '')
                ws[f'H{fila}'] = reg['comprobante'].get('fecha', '')
                ws[f'I{fila}'] = reg['comprobante'].get('hora', '')
                ws[f'J{fila}'] = reg['fecha_creacion']
                fila += 1
        
        self._aplicar_ancho_columnas(ws, len(headers))
    
    def _llenar_hoja_ingresos(self, ws, registros):
        """Llena la hoja de Control de Ingresos"""
        # Headers
        headers = [
            "ID", "Número Acta", "RUC", "Razón Social", 
            "Número Carta", "Recepción", "Vínculo", "Fecha", "Hora", "Fecha Creación"
        ]
        
        self._aplicar_estilo_headers(ws, headers)
        
        # Datos
        fila = 2
        for reg in registros:
            if reg['tipo_informe'] == 'control_ingresos':
                ws[f'A{fila}'] = reg['id']
                ws[f'B{fila}'] = reg['documento'].get('numero_acta', '')
                ws[f'C{fila}'] = reg['documento'].get('ruc', '')
                ws[f'D{fila}'] = reg['documento'].get('nombre_razon_social', '')
                ws[f'E{fila}'] = reg['comprobante'].get('numero_carta', '')
                ws[f'F{fila}'] = reg['comprobante'].get('recepcion', '')
                ws[f'G{fila}'] = reg['comprobante'].get('vinculo', '')
                ws[f'H{fila}'] = reg['comprobante'].get('fecha', '')
                ws[f'I{fila}'] = reg['comprobante'].get('hora', '')
                ws[f'J{fila}'] = reg['fecha_creacion']
                fila += 1
        
        self._aplicar_ancho_columnas(ws, len(headers))
    
    def _aplicar_estilo_headers(self, ws, headers):
        """Aplica estilo a los headers"""
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            celda = ws.cell(row=1, column=col)
            celda.value = header
            celda.fill = fill
            celda.font = font
            celda.alignment = alignment
    
    def _aplicar_ancho_columnas(self, ws, num_columnas):
        """Ajusta ancho de columnas"""
        ancho_default = 15
        for col in range(1, num_columnas + 1):
            ws.column_dimensions[chr(64 + col)].width = ancho_default
